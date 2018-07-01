# https://stackoverflow.com/questions/37737538/merge-matplotlib-subplots-with-shared-x-axis

from openpyxl import Workbook, load_workbook
from matplotlib import ticker, gridspec, pyplot as plt
import matplotlib.lines as mlines
import sys, platform
from IPython import embed as IP
import pandas as pd

if platform.system() == 'Linux':
    sys.path.append("/afs/unity.ncsu.edu/users/b/bli6/temoa/temoa_model")
    sys.path.append('/opt/ibm/ILOG/CPLEX_Studio1263/cplex/python/2.6/x86-64_linux')
elif platform.system() == 'Windows':
    sys.path.append('C:\\Users\\bli\\GitHub\\Temoa\\temoa_model')
    sys.path.append('C:\\Program Files\\IBM\\ILOG\\CPLEX_Studio128\\cplex\\python\\2.7\\x64_win64')
    sys.path.append('C:\\Users\\bxl180002\\git\\Tools')
elif platform.system() == 'Darwin':
    sys.path.append('/Users/bli/git/temoa/temoa_model')
    sys.path.append('/Users/bli/Applications/IBM/ILOG/CPLEX_Studio1263/cplex/python/2.7/x86-64_osx')
    sys.path.append('/Users/bli/git/tools')
else:
    print 'Unrecognized system! Exiting...'
    sys.exit(0)

from plot_result import plot_NCdemand_all, plot_emis_all, TemoaNCResult, plot_bar, plot_load_ts
from mapping import *

def plot_breakeven(ax, years, scenarios, bic, ic, i_subplot=None):
    # bic is a dictionary, ic is a list of the raw investment costs
    # ic = [x, x, ..., x], the length of which equals to the length of years
    # bic[scenario] = [x, x, x... x] where the length equals to the number 
    # of optimized periods.

    handles = list()

    ymax = max(ic)
    for s in scenarios:
        h, = ax.plot(years, bic[s], 
            color = sen_color_map[s],
            marker = sen_marker_map[s],
            linestyle = sen_lstyle_map[s],
            markersize = 2,
            linewidth = 1,
        )
        handles.append(h)

    h = ax.fill_between(years, 0, ic, 
        facecolor = sen_color_map['IC']
        )
    handles.append(h)

    ax.set_xlim( ( years[0]-2, years[-1]+2 ) )
    ymax = int(ymax)
    ymax = ymax - ymax%100 + 100
    ax.set_ylim(0, ymax)
    ax.set_xticks(years)
    if i_subplot:
        ax.annotate(
            i_subplot,
            xy=(0.1, 0.9), xycoords='axes fraction',
            xytext=(.1, .9), textcoords='axes fraction',
            fontsize=12,
                )
    return handles

def plot_6bar_clustered(ax, periods, colortypes, scenarios, values):
    handles = list()
    w     = 0.1*5
    gap   = 0.04*5 # w + gap = 5*(1/7)
    ymax  = 0 # To determine ylimits

    for k in range(0, len(scenarios)):
        b = [0]*len(periods) # Bottom height
        s = scenarios[k]
        offset_bar0 = -(3*w + 2.5*gap) + k*(w + gap)
        x = [i + offset_bar0 for i in periods]
        for t in colortypes:
            y = values[t][s]
            if hatch_map[ category_map[t] ]:
                h = ax.bar(
                    x,
                    y,
                    width = w,
                    bottom = b,
                    color = color_map[ category_map[t] ], 
                    hatch = hatch_map[ category_map[t] ],
                )
            else:
                h = ax.bar(
                    x, 
                    y, 
                    width = w,
                    bottom = b,
                    color = color_map[ category_map[t] ], 
                    edgecolor = edge_map[ category_map[t] ],
                )
            b = [b[i] + values[t][s][i] for i in range(0, len(b))]
            handles.append(h)
        if ymax < max(b):
            ymax = max(b)
    ymax = int(ymax)
    if ymax > 100:
        ymax = ymax - ymax%50 + 65
    elif ymax < 100:
        ymax = ymax - ymax%10 + 15
    ax.set_ylim(0, ymax)

    return handles

# wb1 = load_workbook('Graph.EST.xlsx', data_only=True)
# Size of each individual subplot, in inch
hfig = 4
wfig = 10

def plot_panel(db, scenario, title_names = None):
    ns = len(scenario)
    if ns%2 == 1:
        ncols = 2
        nrows = (ns + 1)/ncols
    else:
        ncols = 2
        nrows = ns/ncols
    for figure_type in ['capacities', 'activities']:
        fig = plt.figure(
            figsize=(wfig/2.0*ncols, hfig*3.0/4*nrows),
            dpi=80,
            facecolor='w',
            edgecolor='k',
        )
        gs = gridspec.GridSpec(
            nrows, ncols,
            height_ratios=[1]*nrows,
            width_ratios=[1]*ncols,
            hspace=0,
            wspace=0,
        ) 
        ax = list()
        ymax = 0
        for s in scenario:
            instance = TemoaNCResult(db, s)
            sindex = scenario.index(s)
            nr = (sindex - (sindex%2))/2
            nc = sindex%2
            ax.append( plt.subplot( gs[nr, nc]) )
            techs  = instance.techs
            values = getattr( instance, figure_type )
            if figure_type == 'activities':
                for t in values:
                    values[t] = [i/3.6 for i in values[t]]
            else:
                techs.remove('EE')
            handles = plot_bar(ax[sindex], instance.periods, instance.techs, values)
            if title_names:
                ax[sindex].text(
                    .5, .9,
                    ' '.join([
                        '(' + chr(ord('a')+sindex) + ')',
                        title_names[sindex]
                    ]),
                    horizontalalignment='center',
                    transform=ax[sindex].transAxes,
                    fontsize=12,
                )
            ylim = ax[sindex].get_ylim()
            if ymax < ylim[-1]:
                ymax = ylim[-1]
        ymax = ymax - ymax%10 + 15
        for s in scenario:
            sindex = scenario.index(s)
            ax[sindex].set_ylim([0, ymax])
        for i in range( 0, ncols*(nrows - 1) ):
            plt.setp(ax[i].get_xticklabels(), visible=False)
        # plt.subplots_adjust(hspace=.0)
        for i in range(1, ns, 2):
            plt.setp(ax[i].get_yticklabels(), visible=False)
        # plt.subplots_adjust(wspace=.0)
        plt.subplots_adjust(left=0.10, right=0.9, top=0.99)
        for i in range( ncols*(nrows-1), ns ):
            ax[i].set_xlabel('Year')
        for i in range(0, ns, 2):
            if figure_type == 'capacities':
                ax[i].set_ylabel('Capacity (GW)')
            else:
                ax[i].set_ylabel('Electricity production (TWh)')
        if nrows >= 2:
            for i in range(1, nrows):
                yticks = ax[i*ncols].yaxis.get_major_ticks()
                yticks[-1].label1.set_visible(False)

        techs_full = list()
        for t in techs:    
            t_full = [k for k, v in category_map.items() if v == t][0]
            techs_full.append(t_full)
        fig.legend(
            handles, techs_full, 
            loc='upper center', 
            ncol=(len(handles)+len(handles)%2)/2, 
            bbox_to_anchor=(0.5, 0.065),
            edgecolor=None
        )

def plot_2panel(db):
    for figure_type in ['capacities', 'activities']:
        fig = plt.figure(figsize=(wfig, hfig+1), dpi=80, facecolor='w', edgecolor='k')
        gs = gridspec.GridSpec(1, 2, height_ratios=[1], width_ratios=[1, 1], hspace=0, wspace=0) 
        ax = list()
        scenario = ['newRPS', 'newRPS_ITC_EWNDON'] # First left, then down
        ymax = 0
        for s in scenario:
            instance = TemoaNCResult(db, s)
            sindex = scenario.index(s)
            nr = (sindex - (sindex%2))/2
            nc = sindex%2
            ax.append( plt.subplot( gs[nr, nc]) )
            techs  = instance.techs
            techs.remove('EE')
            values = getattr( instance, figure_type )
            if figure_type == 'activities':
                for t in values:
                    values[t] = [i/3.6 for i in values[t]]
            handles = plot_bar(ax[sindex], instance.periods, instance.techs, values)
            ax[sindex].text(
                .5, .9, '('+ chr(ord('a')+sindex) +')',
                horizontalalignment='center',
                transform=ax[sindex].transAxes
            )
            box = ax[sindex].get_position()
            ax[sindex].set_position([box.x0, box.y0 + box.height * 0.15,
                 box.width, box.height * 0.95])
            ylim = ax[sindex].get_ylim()
            if ymax < ylim[-1]:
                ymax = ylim[-1]
        ymax = ymax - ymax%10 + 15
        for s in scenario:
            sindex = scenario.index(s)
            ax[sindex].set_ylim([0, ymax])
        # for i in range(0, 2):
        #     plt.setp(ax[i].get_xticklabels(), visible=False)
        # plt.subplots_adjust(hspace=.0)
        for i in range(1, 2, 2):
            plt.setp(ax[i].get_yticklabels(), visible=False)
        # plt.subplots_adjust(wspace=.0)
        # plt.subplots_adjust(left=0.10, right=0.9, top=0.99)
        for i in range( 0, 2 ):
            ax[i].set_xlabel('Year')
        for i in range(0, 2, 2):
            if figure_type == 'capacities':
                ax[i].set_ylabel('Capacity (GW)')
            else:
                ax[i].set_ylabel('Generation (TWh)')
        techs_full = list()
        for t in techs:    
            t_full = [k for k, v in category_map.items() if v == t][0]
            techs_full.append(t_full)
        fig.legend(
            handles, techs_full, 
            loc='upper center', 
            ncol=5, 
            bbox_to_anchor=(0.5, 0.15),
            edgecolor=None
        )

def plot_6breakeven(df_data):
    techs       = ['EWNDON', 'EWNDOFS', 'EURNALWR15', 'EURNSMR', 'ESOLPVDIS', 'EBIOIGCC', ]
    techs_range = ['EWNDON', 'EWNDOFS', 'EURNALWR15', 'EURNSMR', 'ESOLPVDIS', 'EBIOIGCC', ]
    tech_name   = [
        'Onshore wind',
        'Offshore wind',
        'Light water reactor',
        'Small modular reactor',
        'Residential PV',
        'Biomass IGCC',
    ]

    years = range(2015, 2055, 5)
    scenarios_original = ['L', 'R', 'H', 'cap-L', 'cap-R', 'cap-H', 'newRPS']
    scenarios_range = scenarios_original + ['Ldec', 'Linc', 'cap-Hdec', 'cap-Hinc']

    fig  = plt.figure(figsize=(wfig, hfig*3), dpi=80, facecolor='w', edgecolor='k')
    gs = gridspec.GridSpec(3, 2, wspace=0.0, hspace=0.0) 
    ax = list()
    handles = None
    for it in range(0, len(techs)):
        t = techs[it]
        nr = (it - (it%2))/2
        nc = it%2
        ax.append( plt.subplot(gs[nr, nc]) )
        ic  = list()
        bic = dict()
        if t in techs_range:
            scenarios = scenarios_range
        else:
            scenarios = scenarios_original
        for this_s in scenarios:
            # data_this_t = list()
            data_tmp = df_data['BE IC'][
                (df_data['scenario']==this_s)
                & (df_data['technology']==t)
            ]
            data_tmp.reset_index(drop=True, inplace=True)
            bic[this_s] = list( data_tmp )[0: 8]
        ic = list( df_data['IC'][
            (df_data['scenario']=='R')
            & (df_data['technology']==t)
        ])[0: 8]

        title_subplot = ' '.join([
            '('+chr(ord('a') + it)+')',
            tech_name[it]
        ])
        handles = plot_breakeven(ax[it], years, scenarios_original, bic, ic, title_subplot)
        if t in techs_range:
            c = ['g', 'k']
            for s_lb, s_ub in [
                ('Ldec', 'Linc'),
                ('cap-Hdec', 'cap-Hinc'),
            ]:
                lb = bic[s_lb]
                ub = bic[s_ub]
                ax[it].fill_between(years, lb, ub, color=c.pop(), alpha=0.2)
    
    for i in range(0, len(ax)):
        ylim = ax[i].get_ylim()
        ymax = ylim[1] - ylim[1]%500+800
        ax[i].set_ylim([0, ymax])
        ax[i].tick_params(axis="x", direction='in')
        if i%2 == 1:
            ax[i].yaxis.tick_right()
        if i%2 == 0:
            ax[i].set_ylabel('Capital cost ($/kW)')
        if i<len(ax)-2:
            plt.setp(ax[i].get_xticklabels(), visible=False)
        else:
            for tick in ax[i].get_xticklabels():
                # tick.set_rotation(90)
                pass
    names_on_figure = {
        'L':      'L',
        'R':      'R',
        'H':      'H',
        'cap-L':  'Cap-L',
        'cap-R':  'Cap-R',
        'cap-H':  'Cap-H', 
        'newRPS': 'REPS-R'
    }
    legend_names = list()
    for i in scenarios_original:
        legend_names.append(names_on_figure[i])
    fig.legend(
        handles, legend_names+['Capex'], 
        loc='upper center', 
        ncol=8, 
        bbox_to_anchor=(0.5, 0.08),
        edgecolor=None
    )
    # https://stackoverflow.com/q/18619880/4626354
    plt.subplots_adjust(left=0.10, right=0.9, top=0.99)
    return fig

def plot_abatement(db):
    ex_2015_to_2018 = 1.06 # https://www.officialdata.org/2015-dollars-in-2018?amount=1
    SCC = {
        r'SC-$\mathregular{CO}_2$ 2.5%': [64, 71, 78, 84, 90, 97, 102, 109, ],
        r'SC-$\mathregular{CO}_2$ 3%':   [41, 48, 53, 58, 63, 69, 74,  79, ],
        r'SC-$\mathregular{CO}_2$ 5%':   [13, 14, 16, 18, 21, 24, 26,  30, ],
    }
    lstyle = {
        r'SC-$\mathregular{CO}_2$ 2.5%': '-.',
        r'SC-$\mathregular{CO}_2$ 3%':   '-',
        r'SC-$\mathregular{CO}_2$ 5%':   '--',
    }
    periods     = None
    cap_on      = ['capL', 'capR', 'capH', 'newRPS', 'newRPS_ITC_EWNDON']
    cap_off     = ['L',    'R',    'H',    'R',      'R']
    xlabelnames = cap_on
    cost        = dict()
    co2_emis    = dict()
    ymax        = 0
    for s in cap_off+cap_on:
        instance = TemoaNCResult(db, s)
        periods  = instance.periods
        cost[s]  = instance.TotalCost
        co2_emis[s] = sum(i for i in instance.emissions['co2_ELC'])
    
    delta_cost = list()
    delta_emis = list()
    abatement  = list()
    ITC_credit = [0, 0, 0, 0, 73*ex_2015_to_2018]
    # At present it is hardwired, but in the future we should let the script 
    # to calculate it
    for i in range(0, len(cap_on)):
        s_on, s_off = cap_on[i], cap_off[i]
        delta_cost.append(cost[s_on] - cost[s_off])
        delta_emis.append(co2_emis[s_off] - co2_emis[s_on])
        abatement.append(delta_cost[i]/delta_emis[i]*1000*ex_2015_to_2018)
    abatement.append(0)
    ITC_credit.append(0)
    SCC5_b = [0, 0, 0, 0, 0, min(SCC[r'SC-$\mathregular{CO}_2$ 5%'])*ex_2015_to_2018 ]
    SCC5_t = [0, 0, 0, 0, 0, max(SCC[r'SC-$\mathregular{CO}_2$ 5%'])*ex_2015_to_2018 ]

    # fig  = plt.figure(figsize=(wfig, hfig), dpi=80, facecolor='w', edgecolor='k')
    # gs   = gridspec.GridSpec(1, 2, height_ratios=[1], width_ratios=[1, 1], hspace=0, wspace=0)
    axes = list()
    handles = list()
    legends = list()

    # CO2 abatement cost
    # xlabelnames.append(r'SC-$\mathregular{CO}_2$ 5%')
    xlabelnames = ['Cap-L', 'Cap-R', 'Cap-H', 'REPS-R', 'REPS-R-ITC', ' ']
    fig = plt.figure()
    ax = plt.subplot()
    axes.append(ax)
    h = ax.bar(
        xlabelnames[0: -1],
        abatement[0: -1],
        facecolor = 'white',
        edgecolor = 'black',
    )
    legends.append('abatement cost')
    handles.append(h)
    h = ax.bar(
        xlabelnames[0: -1],
        ITC_credit[0: -1],
        bottom=abatement[0: -1],
        facecolor = 'black',
        edgecolor = 'black',
    )
    # box = ax.get_position()
    # ax.set_position([box.x0, box.y0 + box.height * 0.15,
    #         box.width, box.height * 0.95])
    handles.append(h)
    legends.append('ITC')
    h = ax.bar(
        xlabelnames,
        SCC5_t,
        bottom=SCC5_b,
        facecolor = [0.5, 0.5, 0.5],
        edgecolor = [0.5, 0.5, 0.5],
    )
    legends.append(r'SC-$\mathregular{CO}_2$, 5% discount rate')
    handles.append(h)
    ylim = ax.get_ylim()
    if ymax < ylim[-1]:
        ymax = ylim[-1]
    ax.tick_params(axis='x', length=0)
    tmp_ylabel = r'$\mathregular{CO}_2$ abatement cost (' + '\$/tonne ' + r'$\mathregular{CO}_2$)'
    ax.set_ylabel(tmp_ylabel)
    barnames = ['cap-L', 'cap-R', 'cap-H', 'REPS-R', 'REPS-R-ITC']
    # ax.text(
    #         0.09, 0.5, 'cap-L', transform=ax.transAxes,
    #     )
    # ax.text(
    #         0.27, 0.45, 'cap-R', transform=ax.transAxes,
    #     )
    # ax.text(
    #         0.45, 0.35, 'cap-H', transform=ax.transAxes,
    #     )
    # ax.text(
    #         0.61, 0.30, 'newREPS', transform=ax.transAxes,
    #     )
    # ax.text(
    #         0.81, 0.8, 'newREPS\n+ITC', transform=ax.transAxes,
    #     )
    ax.text(
        0.84, 0.5, r'SC-$\mathregular{CO}_2$', transform=ax.transAxes,
    )
    plt.legend(
        handles, legends, 
        loc='upper left', 
        # ncol=5, 
        # bbox_to_anchor=(0.5, 0.07),
        # edgecolor=None
    )

    # SCC CO2 cost
    fig = plt.figure(0)
    ax = plt.subplot()
    axes.append(ax)
    # handles = list()
    # legends = list()
    for rate in SCC:
        h = ax.plot(
            periods,
            SCC[rate],
            color='k',
            marker='s',
            linestyle=lstyle[rate],
        )
        h = mlines.Line2D([], [], color='k', marker='s', linestyle = lstyle[rate])
        handles.append(h)
        legends.append(rate)
        box = ax.get_position()
        ax.set_position([box.x0, box.y0 + box.height * 0.15,
                box.width, box.height * 0.95])
    ylim = ax.get_ylim()
    if ymax < ylim[-1]:
        ymax = ylim[-1]
    ax.yaxis.tick_right()
    tmp_ylabel = r'Social cost of $\mathregular{CO}_2$ (' + '\$/tonne ' + r'$\mathregular{CO}_2$)'
    ax.set_ylabel(tmp_ylabel)
    ax.yaxis.set_label_position("right")

    # for ax in axes:
    #     ax.set_ylim([0, ymax])
    #     ax.text(
    #                 .5, .9, '('+ chr(ord('b')+axes.index(ax)) +')',
    #                 horizontalalignment='center',
    #                 transform=ax.transAxes
    #             )
    # plt.subplots_adjust(left=0.10, right=0.9, top=0.99)
    # plt.setp(axes[0].get_xticklabels(), visible=False)
    # fig.legend(
    #     handles, legends, 
    #     loc='upper center', 
    #     ncol=5, 
    #     bbox_to_anchor=(0.5, 0.07),
    #     edgecolor=None
    # )

def plot_6load_ts(db, remove_spike=False):
    names_on_figure = {
        'L':      'L',
        'R':      'R',
        'H':      'H',
        'capL':   'Cap-L',
        'capR':   'Cap-R',
        'capH':   'Cap-H', 
        'newRPS': 'REPS-R'
    }
    for this_p in range(2015, 2055, 5):
        fig = plt.figure(figsize=(wfig, hfig*3), dpi=80, facecolor='w', edgecolor='k')
        gs = gridspec.GridSpec(
            3, 2,
            height_ratios=[1, 1, 1],
            width_ratios=[1, 1],
            hspace=0.05,
            wspace=0.05,
        ) 
        ax = list()
        scenario = ['L', 'capL', 'R', 'capR', 'H', 'capH'] # First left, then down
        ymax = 0
        for s in scenario:
            instance = TemoaNCResult(db, s)
            sindex = scenario.index(s)
            nr = (sindex - (sindex%2))/2
            nc = sindex%2
            ax.append( plt.subplot( gs[nr, nc]) )
            index_text = '('+ chr(ord('a')+sindex) +')' + ' ' + names_on_figure[s]
            area_handles, area_names = plot_load_ts(ax[sindex], instance, this_p, index_text, remove_spike)
            # ax[sindex].text(
            #     .5, .9, '('+ chr(ord('a')+sindex) +')',
            #     horizontalalignment='center',
            #     transform=ax[sindex].transAxes,
            #     fontsize=12,
            # )
            ylim = ax[sindex].get_ylim()
            if ymax < ylim[-1]:
                ymax = ylim[-1]
        ymax = ymax - ymax%10 + 15
        for s in scenario:
            sindex = scenario.index(s)
            ax[sindex].set_ylim([0, ymax])
        for i in range(0, 4):
            plt.setp(ax[i].get_xticklabels(), visible=False)
        # plt.subplots_adjust(hspace=.0)
        for i in range(1, 6, 2):
            plt.setp(ax[i].get_yticklabels(), visible=False)
        # plt.subplots_adjust(wspace=.0)
        plt.subplots_adjust(left=0.10, right=0.9, top=0.99)
        for i in range( 4, 6 ):
            ax[i].set_xlabel('Time slice')
        for i in range(0, 5, 2):
            ax[i].set_ylabel('Average power (GW)')

        fig.legend(
            area_handles, area_names, 
            loc='upper center', 
            ncol=(len(area_names) + len(area_names)%2)/2, 
            bbox_to_anchor=(0.5, 0.065),
            edgecolor=None
        )

# The following are old functions for my PhD thesis paper

def cap_and_act(db):
    fig = plt.figure(figsize=(wfig, hfig*2), dpi=80, facecolor='w', edgecolor='k')
    gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1]) 
    ax = list()
    attributes = ['activities', 'capacities']
    for a in attributes:
        ax.append( plt.subplot(gs[attributes.index(a)]) )
        scenarios  = ['L', 'R', 'H', 'capL', 'capR', 'capH']
        values     = dict()
        periods    = None
        for s in scenarios:
            instance = TemoaNCResult(db, s)
            periods  = instance.periods
            for t in instance.techs:
                t_full = [k for k, v in category_map.items() if v == t][0]
                if t_full not in values:
                    # Inverse lookup, see https://stackoverflow.com/q/2568673
                    values[t_full] = dict()
                values[t_full][s] = getattr(instance, a)[t]
                if a is 'activities':
                    values[t_full][s] = [ i/3.6 for i in values[t_full][s] ]
        technology = values.keys()
        for t in technology:
            for s in scenarios:
                if s not in values[t]:
                    values[t][s] = [0 for i in periods]
        if 'Geothermal' in technology:
            technology.remove('Geothermal')
        if 'EE' in technology:
            technology.remove('EE')
        for t in [
            'Nuclear',      'Hydro',     'Oil',
            'Pumped hydro', 'Coal',      'Natural gas',
            'Wind',         'Solar PV',  'Bioenergy',
        ]:
            if t in technology:
                technology.remove(t)
                technology.append(t)
        handles = plot_6bar_clustered(ax[attributes.index(a)], periods, technology, scenarios, values)

    ax[1].set_xlabel('Periods')
    ax[1].set_ylabel('GW')
    ax[0].set_ylabel('TWh')
    plt.setp(ax[0].get_xticklabels(), visible=False) # shared x axis
    yticks = ax[1].yaxis.get_major_ticks()
    yticks[-1].label1.set_visible(False) # remove last tick label for the second subplot
    plt.subplots_adjust(hspace=.0) # remove vertical gap between subplots
    # ax[1].legend(handles, technology, loc='upper left')
    fig.legend(
        handles, technology, 
        loc='upper center', 
        ncol=5, 
        bbox_to_anchor=(0.5, 0.065),
        edgecolor=None
    )
    # https://stackoverflow.com/q/18619880/4626354
    plt.subplots_adjust(left=0.06, right=0.94, top=0.99, wspace=0, hspace=0)
    return fig

def break_even():
    def return_figure(tech):
        fig  = plt.figure(figsize=(wfig*len(tabs), hfig*len(tech)), dpi=80, facecolor='w', edgecolor='k')
        gs = gridspec.GridSpec(len(tech), len(tabs), wspace=0.05, hspace=0.0) 
        ax = list()
        handles = None
        counter = 0
        for it in range(0, len(tech)):
            t = tech[it]
            for itab in range(0, len(tabs)):
                tab = tabs[itab]
                ax.append( plt.subplot(gs[it*2+itab]) )
                scenarios = ['L', 'R', 'H', 'CPP-L', 'CPP-R', 'CPP-H']
                periods   = range(2015, 2055, 5)
                ws1 = wb1[tab]
                ic  = list()
                bic = dict()
                for j in range(0, len(periods)):
                    v = ws1.cell(row=nr0[t]+j, column=nc0[t]+6).value
                    ic.append(v)

                for i in range(0, len(scenarios)):
                    nc = nc0[t] + i
                    s  = scenarios[i]
                    bic[s] = list()
                    for j in range(0, len(periods)):
                        nr = nr0[t] + j
                        v = ws1.cell(row=nr, column=nc).value
                        bic[s].append(v)
                handles = plot_breakeven(ax[it*2+itab], periods, scenarios, bic, ic, '('+chr(ord('a') + counter)+')')
                counter += 1
        
        for i in range(0, len(ax)):
            if i%2 == 1:
                plt.setp(ax[i].get_yticklabels(), visible=False)
            if i%2 == 0:
                ax[i].set_ylabel('Capital cost ($/MW)')
            if i<len(ax)-2:
                plt.setp(ax[i].get_xticklabels(), visible=False)
            else:
                for tick in ax[i].get_xticklabels():
                    tick.set_rotation(90)

        fig.legend(
            handles, scenarios+['Capex'], 
            loc='upper center', 
            ncol=7, 
            bbox_to_anchor=(0.5, 0.045),
            edgecolor=None
        )
        # https://stackoverflow.com/q/18619880/4626354
        plt.subplots_adjust(left=0.10, right=0.9, top=0.99, wspace=0, hspace=0)
        return fig

    hfig = 2.5
    wfig = 4
    tabs = ['SenAnalysis RPS (CPLEX API)', 'SenAnalysis no RPS (CPLEX API)']
    nr0 = {
        'Solar PV':      1451,
        'Onshore wind':  1461,
        'Biomass IGCC':  1471,
        'Nuclear':       1481,
        'Coal IGCC':     1491, 
        'NG IGCC':       1501,
        'Offshore wind': 1511,
    }
    nc0 = {
        'Solar PV':      3,
        'Onshore wind':  3,
        'Biomass IGCC':  3,
        'Nuclear':       3,
        'Coal IGCC':     3, 
        'NG IGCC':       3,
        'Offshore wind': 3,
    }

    figs = list()
    tech = ['Solar PV', 'Onshore wind', 'Biomass IGCC']
    fig = return_figure(tech)
    figs.append(fig)
    tech = ['Nuclear','Coal IGCC', 'NG IGCC','Offshore wind']
    fig = return_figure(tech)
    figs.append(fig)
    return figs

def emissions(db):
    emissions = ['so2_ELC', 'nox_ELC', 'co2_ELC']
    scenarios = ['L', 'R', 'H', 'capL', 'capR', 'capH']

    fig = plt.figure(figsize=(wfig, hfig*3), dpi=80, facecolor='w', edgecolor='k')
    gs = gridspec.GridSpec(3, 1, height_ratios=[1, 1, 1]) 
    ax = list()
    for ie in range(0, len(emissions)):
        e = emissions[ie]
        ax.append( plt.subplot(gs[ie]) )

        values     = dict()
        periods    = None
        for s in scenarios:
            instance = TemoaNCResult(db, s)
            periods  = instance.periods
            for t in instance.emissions1[e]:
                t_full = [k for k, v in category_map.items() if v == t][0]
                if t_full not in values:
                    values[t_full] = dict()
                values[t_full][s] = instance.emissions1[e][t]
                if e == 'co2_ELC':
                    values[t_full][s] = [
                        i/1000.0 for i in instance.emissions1[e][t]
                    ] # kilotonne to megatonne
                else:
                    values[t_full][s] = instance.emissions1[e][t]
            if 'Emission reduction' not in values:
                values['Emission reduction'] = dict()
            if e == 'co2_ELC':
                values['Emission reduction'][s] = [
                    i/1000.0 for i in instance.emis_redct[e]
                ]
            else:
                values['Emission reduction'][s] = instance.emis_redct[e]

        technology = values.keys()
        if 'other' in technology:
            technology.remove('other')
        if 'Total raw' in technology:
            technology.remove('Total raw')
        for t in ['Oil', 'Bioenergy', 'Natural gas', 'Coal', 'Emission reduction']:
            if t in technology:
                technology.remove(t)
                technology.append(t)
        handles = plot_6bar_clustered(ax[ie], periods, technology, scenarios, values)

    ax[-1].set_xlabel('Periods')
    ax[0].set_ylabel('Emission (kilotonnes)')
    ax[1].set_ylabel('Emission (kilotonnes)')
    ax[2].set_ylabel('Emission (megatonnes)')
    plt.setp(ax[0].get_xticklabels(), visible=False) # shared x axis
    plt.setp(ax[1].get_xticklabels(), visible=False) # shared x axis
    yticks = ax[2].yaxis.get_major_ticks()
    yticks[-1].label1.set_visible(False) # remove last tick label for the last subplot
    yticks = ax[1].yaxis.get_major_ticks()
    yticks[-1].label1.set_visible(False) # remove last tick label for the second last subplot
    plt.subplots_adjust(hspace=.0) # remove vertical gap between subplots
    # ax[1].legend(handles, technology, loc='upper left')
    fig.legend(
        handles, technology, 
        loc='upper center', 
        ncol=5, 
        bbox_to_anchor=(0.5, 0.065),
        edgecolor=None
    )
    # https://stackoverflow.com/q/18619880/4626354
    plt.subplots_adjust(left=0.06, right=0.94, top=0.99, wspace=0, hspace=0)
    return fig

def plot_marginalCO2():
    fig = plt.figure(figsize=(5, 3), dpi=80, facecolor='w', edgecolor='k')
    ax  = plt.subplot(111)
    w = 1
    gap = 2
    years = range(2025, 2055, 5)
    values = dict()
    values['CPP-L']         = [0.00, 4.68, 0.00, 51.12, 53.86, 28.89]
    values['CPP-R']         = [17.88, 33.40, 23.92, 50.79, 41.97, 27.05]
    values['CPP-H']         = [86.17, 81.90, 57.81, 58.51, 47.70, 27.61]
    values['SCC in 2015 $'] = [16, 18, 21, 24, 26, 30]

    scenarios = ['CPP-L', 'CPP-R', 'CPP-H']
    fcolor    = ['k', [0.8, 0.8, 0.8],  'w']
    ecolor    = ['k', [0.8, 0.8, 0.8],  'k']
    handles   = list()
    for i in range(0, len(scenarios)):
        s = scenarios[i]
        offset = (i - 1)*w
        h = ax.bar(
            [j + offset for j in years],
            values[s],
            width=w,
            color=fcolor[i],
            edgecolor=ecolor[i],
        )
        handles.append(h)
    h, = ax.plot(
        years,
        values['SCC in 2015 $'],
        linewidth=1,
        color='k',
        marker='s',
        markersize=2,
    )
    handles.append(h)
    plt.legend(
        handles,
        scenarios+['SCC in 2015 $'],
        loc='upper center', 
        ncol=4, 
        bbox_to_anchor=(0.5, 1.15),
        edgecolor=None
    )
    ax.set_ylabel(r'\$/tonne $\mathregular{CO}_2$')
    return fig

def CF_solar_wind_load():
    tech = ['Onshore wind', 'Offshore wind', 'Solar PV, rooftop', 'Solar PV, utility']
    color_tech = {
        'Onshore wind':      'g',
        'Offshore wind':     'g',
        'Solar PV, rooftop': 'r',
        'Solar PV, utility': 'r',
        'Hourly load'      : 'k',
    }
    lstyle_tech = {
        'Onshore wind':      '--',
        'Offshore wind':     '-',
        'Solar PV, rooftop': '--',
        'Solar PV, utility': '-',
        'Hourly load'      : '-',

    }
    seg  = range(1, 97)
    wb = load_workbook(
        'LDC.xlsx',
        data_only=True,
    )
    nr0 = {
        'Onshore wind':       7,
        'Offshore wind':      7,
        'Solar PV, rooftop':  7,
        'Solar PV, utility':  7,
    }
    nc0 = {
        'Onshore wind':       7,
        'Offshore wind':      8,
        'Solar PV, rooftop':  10,
        'Solar PV, utility':  11,
    }
    ws = wb['CF_summary']
    values = dict()
    for j in range(0, len(tech)):
        t = tech[j]
        values[t] = list()
        for i in range(0, len(seg)):
            v = ws.cell(row=nr0[t]+i, column=nc0[t]).value
            values[t].append(v)

    ws = wb['By Seasons']
    nr0 = 83
    nc0 = 7
    values['Hourly load'] = list()
    for j in range(0, len(seg)):
        v = ws.cell(row=nr0, column=nc0+j).value
        values['Hourly load'].append(v)

    fig = plt.figure(figsize=(7, 4), dpi=80, facecolor='w', edgecolor='w')
    ax1 = plt.subplot(111)
    handles = list()
    for t in tech:
        h, = ax1.plot(
            seg,
            values[t],
            linewidth=1,
            color=color_tech[t],
            linestyle=lstyle_tech[t],
        )
        handles.append(h)
    xticks      = [1, 25, 49, 73]
    yticks      = [i*0.1 for i in range(0, 7)]
    vals        = ax1.get_yticks()
    ax1.set_yticklabels(['{:d}%'.format(int(x*100)) for x in vals])
    ax1.xaxis.set_major_locator(ticker.FixedLocator(xticks))
    ax1.set_xlabel('Time Slice')
    ax1.set_ylabel('Availability factor')

    ax2 = ax1.twinx()
    h, = ax2.plot(
        seg,
        [i/1E3 for i in values['Hourly load']],
        linewidth=1,
        color=color_tech['Hourly load'],
        linestyle=lstyle_tech['Hourly load'],
    )
    handles.append(h)
    ax2.set_ylabel('Load (GW)')
    plt.legend(
        handles,
        tech + ['Hourly load'],
        loc = 'upper center',
        bbox_to_anchor=[0.5, 1.17],
        ncol=3,
    )
    return fig