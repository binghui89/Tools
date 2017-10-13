# This version is compatible with Pyomo 5.2

import os, sys
from pyomo.environ import *
from pyomo.pysp.scenariotree.manager import \
    ScenarioTreeManagerClientSerial
from pyomo.pysp.ef import create_ef_instance
from pyomo.opt import SolverFactory
from time import time

from IPython import embed as IP


# To see detailed information about options
#for name in options.keys():
#    print(options.about(name))

# To see a more compact display of options
#options.display()

# options.model_location = \
#     os.path.join(farmer_example_dir, 'models')
# options.scenario_tree_location = \
#     os.path.join(farmer_example_dir, 'scenariodata')

class DummyTemoaConfig():
    pass

def return_CP_and_path(p_data):
    # return_CP_and_path(p_data) -> dict(), dict()
    # This function reads the path to the instance directory (p_data) and 
    # returns conditional two dictionaries, the first one is the conditional 
    # probability condition of a scenario, the second one is the path to
    # all files of a scenario.
    from collections import deque, defaultdict
    from pyomo.pysp.util.scenariomodels import scenario_tree_model
    pwd = os.getcwd()
    os.chdir(p_data)

    s2fp_dict = defaultdict(deque) # Scenario to 'file path' dictionary, .dat not included
    s2cd_dict = defaultdict(float) # Scenario to conditonal density mapping
    sStructure = scenario_tree_model.create_instance( filename='ScenarioStructure.dat' )

    # The following code is borrowed from Kevin's temoa_lib.py
    ###########################################################################
    # Step 1: find the root node.  PySP doesn't make this very easy ...

    # a child -> parent mapping, because every child has only one parent, but
    # not vice-versa
    ctpTree = dict() # Child to parent dict, one to one mapping

    to_process = deque()
    to_process.extend( sStructure.Children.keys() )
    while to_process:
        node = to_process.pop()
        if node in sStructure.Children:
            # it's a parent!
            new_nodes = set( sStructure.Children[ node ] )
            to_process.extend( new_nodes )
            ctpTree.update({n : node for n in new_nodes })

                     # parents           -     children
    root_node = (set( ctpTree.values() ) - set( ctpTree.keys() )).pop()

    # ptcTree = defaultdict( list ) # Parent to child node, one to multiple mapping
    # for c, p in ctpTree.iteritems():
    #         ptcTree[ p ].append( c )
    # ptcTree = dict( ptcTree )   # be slightly defensive; catch any additions

    # leaf_nodes = set(ctpTree.keys()) - set(ctpTree.values())
    leaf_nodes = set(sStructure.ScenarioLeafNode.values()) # Try to hack Kevin's code
    
    scenario_nodes = dict() # Map from leafnode to 'node path'
    for node in leaf_nodes: # e.g.: {Rs0s0: [R, Rs0, Rs0s0]}
        s = deque()
        scenario_nodes[ node ] = s
        while node in ctpTree:
            s.append( node )
            node = ctpTree[ node ]
        s.append( node )
        s.reverse()
    ###########################################################################

    for s in sStructure.Scenarios:
        cp = 1.0 # Starting probability
        for n in scenario_nodes[sStructure.ScenarioLeafNode[s]]:
            cp = cp*sStructure.ConditionalProbability[n]
            if not sStructure.ScenarioBasedData.value:
                s2fp_dict[s].append(n + '.dat')
        s2cd_dict[s] = cp
    
    from pyomo.core import Objective
    if sStructure.ScenarioBasedData.value:
        for s in sStructure.Scenarios:
            s2fp_dict[s].append(s + '.dat')
    os.chdir(pwd)
    return (s2cd_dict, s2fp_dict)

def return_breakeven(ef_instance, instance, techs):
    # Return break-even costs for stochastic runs. Need the extensive form
    # instance (ef_instance), and scenario-specific instance (instance), and the
    # technologies of which the break-even costs are required.
    coef_CAP = dict()
    scal_CAP = dict()
    # Break-even investment cost for this scenario, indexed by technology
    years    = list()
    bic_s    = dict()
    ic_s     = dict() # Raw investment costs for this scenario, indexed by tech
    for t in techs:
        vintages = instance.vintage_optimize
        P_0 = min( instance.time_optimize )
        GDR = value( instance.GlobalDiscountRate )
        MLL = instance.ModelLoanLife
        MPL = instance.ModelProcessLife
        LLN = instance.LifetimeLoanProcess
        x   = 1 + GDR    # convenience variable, nothing more.

        bic_s[t] = list()
        ic_s[t]  = list()
        years = vintages.value
        for v in vintages:
            if (t, v) not in instance.CostInvest:
                continue
            period_available = set()
            for p in instance.time_future:
                if (p, t, v) in instance.CostFixed.keys():
                    period_available.add(p)

            # Coefficient for investment cost
            c_i = ( 
                    instance.CostInvest[t, v] 
                    * instance.LoanAnnualize[t, v] 
                    * ( LLN[t, v] if not GDR else 
                        (x**(P_0 - v + 1) 
                        * ( 1 - x **( -value(LLN[t, v]) ) ) 
                        / GDR) 
                      ) 
            )

            # Coefficient for salvage value
            c_s = (-1)*(
                value( instance.CostInvest[t, v] )
                * value( instance.SalvageRate[t, v] )
                / ( 1 if not GDR else 
                    (1 + GDR)**( 
                        instance.time_future.last() 
                        - instance.time_future.first()
                        - 1
                        ) 
                    )
                )

            # Coefficient for fixed cost
            c_f = sum( 
                instance.CostFixed[p, t, v]
                * ( MPL[p, t, v] if not GDR else
                    (x**(P_0 - p + 1)
                    * ( 1 - x**( -value(MPL[p, t, v]) ) )
                    / GDR ) 
                  )
                for p in period_available 
            ) 

            c = c_i + c_s + c_f
            s = (c - ef_instance.lrc[instance.V_Capacity[t, v]])/c
            coef_CAP[t, v] = c
            scal_CAP[t, v] = s # Must reduce TO this percentage
            bic_s[t].append(scal_CAP[t, v]*instance.CostInvest[t, v])
            ic_s[t].append(instance.CostInvest[t, v])

        # print "{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>7s}\t{:>5s}\t{:>3s}\t{:>5s}".format('Tech','Vintage', 'L. RC', 'Coef', 'U. RC', 'Scale', 'BE IC', 'BE FC', 'IC', 'FC', 'Cap')
        print "{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>5s}\t{:>5s}".format('Tech','Vintage', 'L. RC', 'Coef', 'U. RC', 'Scale', 'BE IC', 'IC', 'Cap')
        for v in vintages:
            if (t, v) not in instance.CostInvest:
                continue
            lrc = ef_instance.lrc[instance.V_Capacity[t, v]]
            urc = ef_instance.urc[instance.V_Capacity[t, v]]

            # print "{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>7.1f}\t{:>5.0f}\t{:>3.0f}\t{:>5.3f}".format(
            # t, v, lrc, coef_CAP[t, v], urc, scal_CAP[t, v], 
            # scal_CAP[t, v]*instance.CostInvest[t, v], 
            # scal_CAP[t, v]*instance.CostFixed[v, t, v], # Use the FC of the first period
            # instance.CostInvest[t,v],
            # instance.CostFixed[v, t, v],
            # value(instance.V_Capacity[t, v])
            # )
            print "{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>5.0f}\t{:>5.3f}".format(
            t, v, lrc, coef_CAP[t, v], urc, scal_CAP[t, v], 
            scal_CAP[t, v]*instance.CostInvest[t, v], 
            instance.CostInvest[t,v],
            value(instance.V_Capacity[t, v])
            )

    # print 'Dual and slack variables for emission caps:'
    # for e in instance.commodity_emissions:
    #     for p in instance.time_optimize:
    #         if (p, e) in instance.EmissionLimitConstraint:
    #             print p, e, ef_instance.dual[instance.EmissionLimitConstraint[p, e]], '\t', ef_instance.slack[instance.EmissionLimitConstraint[p, e]]
    return years, bic_s, ic_s

def compute_evpi(ef_result, pf_result):
    pf = 0
    for i in range( 0, len(pf_result['cost']) ):
        pf += pf_result['cd'][i]*pf_result['cost'][i]
    return ef_result - pf

def solve_pf(p_model, p_data):
    """
    solve_pf(p_model, p_data) -> dict()
    Solves the model in perfect sight mode. 
    p_model -> string, the path to the model file. 
    p_data -> string, the path to the directory of data for the stochastic
    mdoel, where ScenarioStructure.dat should resides.
    Returns a dictionary including the value of objective function for each
    scenario and its conditional probability.
    """

    def return_obj(instance):
        from pyomo.core import Objective
        obj = instance.component_objects(Objective, active = True)
        obj_values = list()
        for o in obj:
            # See section 18.6.3 in Pyomo online doc
            # https://stackoverflow.com/a/3071/4626354
            method_obj = getattr(instance, str(o))
            obj_values.append(method_obj())
        # Assuming there is only one objective function
        return obj_values[0]

        # Out-of-date for Pyomo 4.1
        # obj = instance.active_components(Objective) 
        # objs = obj.items()[0]
        # obj_name, obj_value = objs[0], value(objs[1]())
        # return obj_value

    from pyomo.core import Objective

    (head, tail) = os.path.split(p_model)
    sys.path.insert(0, head)

    s2cd_dict, s2fp_dict = return_CP_and_path(p_data)

    pwd = os.getcwd()
    os.chdir(p_data)
    model_module = __import__(tail[:-3], globals(), locals())
    model = model_module.model
    pf_result = {'cost': list(), 'cd': list()}
    # for s in sStructure.Scenarios:
    for s in s2fp_dict:
        pf_result['cd'].append(s2cd_dict[s])
        data = DataPortal(model=model)
        for dat in s2fp_dict[s]:
            data.load(filename=dat)
        instance = model.create_instance(data)
        optimizer = SolverFactory('cplex')
        results = optimizer.solve(instance)

        instance.solutions.load_from(results)
        # instance.load(results)
        obj_val = return_obj(instance)
        pf_result['cost'].append(obj_val)
        sys.stdout.write('\nSolved .dat(s) {}\n'.format(s2fp_dict[s]))
        sys.stdout.write('    Total cost: {}\n'.format(obj_val))
    os.chdir(pwd)
    return pf_result

def solve_ef(p_model, p_data, dummy_temoa_options = None):
    """
    solve_ef(p_model, p_data) -> objective value of the extensive form
    Solves the model in stochastic mode. 
    p_model -> string, the path to the model file (ReferenceModel.py). 
    p_data -> string, the path to the directory of data for the stochastic
    mdoel, where ScenarioStructure.dat should resides.
    Returns a float point number of the value of objective function for the
    stochastic program model.
    """

    options = ScenarioTreeManagerClientSerial.register_options()

    if os.path.basename(p_model) == 'ReferenceModel.py':
        options.model_location = os.path.dirname(p_model)
    else:
        sys.stderr.write('\nModel file should be ReferenceModel.py. Exiting...\n')
        sys.exit(1)
    options.scenario_tree_location = p_data

    # using the 'with' block will automatically call
    # manager.close() and gracefully shutdown
    with ScenarioTreeManagerClientSerial(options) as manager:
        manager.initialize()

        ef_instance = create_ef_instance(manager.scenario_tree,
                                         verbose_output=options.verbose)

        ef_instance.dual  = Suffix(direction=Suffix.IMPORT)
        ef_instance.lrc   = Suffix(direction=Suffix.IMPORT)
        ef_instance.urc   = Suffix(direction=Suffix.IMPORT)
        ef_instance.slack = Suffix(direction=Suffix.IMPORT)

        with SolverFactory('cplex') as opt:

            ef_result = opt.solve(
                ef_instance,
                suffixes=['dual', 'urc', 'slack', 'lrc']
            )

        # Write to database
        if dummy_temoa_options:
            sys.path.append(options.model_location)
            from pformat_results import pformat_results
            from temoa_config import TemoaConfig
            temoa_options = TemoaConfig()
            temoa_options.config = dummy_temoa_options.config
            temoa_options.keepPyomoLP = dummy_temoa_options.keepPyomoLP
            temoa_options.saveTEXTFILE = dummy_temoa_options.saveTEXTFILE
            temoa_options.path_to_db_io = dummy_temoa_options.path_to_db_io
            temoa_options.saveEXCEL = dummy_temoa_options.saveEXCEL
            ef_result.solution.Status = 'feasible' # Assume it is feasible
            # Maybe there is a better solution using manager, but now it is a 
            # kludge to use return_CP_and_path() function
            s2cd_dict, s2fp_dict = return_CP_and_path(p_data)
            for s in manager.scenario_tree.scenarios:
                ins = s._instance
                temoa_options.scenario = s.name
                temoa_options.dot_dat = list()
                for fname in s2fp_dict[s.name]:
                    temoa_options.dot_dat.append(
                        os.path.join(options.scenario_tree_location, fname)
                    )
                temoa_options.output = os.path.join(
                    options.scenario_tree_location, 
                    dummy_temoa_options.output
                    )
                msg = '\nStoring results from scenario {} to database.\n'.format(s.name)
                sys.stderr.write(msg)
                formatted_results = pformat_results( ins, ef_result, temoa_options )

    ef_instance.solutions.store_to( ef_result )
    ef_obj = value( ef_instance.EF_EXPECTED_COST.values()[0] )
    return ef_obj

def do_test(p_model, p_data, temoa_config = None):
    from time import time
    t0 = time()
    timeit = lambda: time() - t0

    if not isinstance(p_data, list):
        p_data = [p_data]
    for this_data in p_data:
        sys.stderr.write('\nSolving perfect sight mode\n')
        sys.stderr.write('-'*25 + '\n')
        pf_result = solve_pf(p_model, this_data)
        msg = 'Time: {} s\n'.format( timeit() )
        sys.stderr.write(msg)
    
        sys.stderr.write('\nSolving extensive form\n')
        sys.stderr.write('-'*25 + '\n')
        ef_result = solve_ef(p_model, this_data, temoa_config)
    
        msg = '\nTime: {} s\n'.format( timeit() )
        sys.stderr.write(msg)
        msg += 'runef objective value: {}\n'.format(ef_result)
        msg += 'EVPI: {}\n'.format( compute_evpi(ef_result, pf_result) )
        sys.stdout.write(msg)

def handle_excel(source, target):
    # Copy data from source and put it into target.
    from openpyxl import load_workbook
    sheets = [
    u'Costs',
    u'Activity_residential',
    u'Emissions',
    u'Capacity_residential',
    u'Activity_electric',
    u'Capacity_electric',
    u'Activity_supply',
    u'Capacity_supply'
    ]

    # wb1 = load_workbook('template.xlsx')
    # ws = wb1['Activity_electric']
    # for row in ws.iter_rows():
    #     pass
    
    # IP()
    wb1 = load_workbook(source)
    wb2 = load_workbook(target)
    for sheet in sheets:
        if sheet == 'Costs':
            pass
        elif sheet == 'Emissions':
            pass
        else:
            print sheet
            ws1 = wb1[sheet]
            ws2 = wb2[sheet]
            name2nrow = dict()
            nrow1 = 1
            for row in ws1.iter_rows():
                name2nrow[ row[0].value ] = nrow1
                nrow1 += 1
            nrow2 = 1
            for row in ws2.iter_rows():
                nrow2 += 1
                if nrow2 == 1:
                    continue
                else:
                    for c in row:
                        if c.col_idx == 1:
                            name = c.value
                            nrow1 = name2nrow[ name ]
                        else:
                            c.value = ws1.cell(
                                row = nrow1, 
                                column = c.col_idx
                                ).value
                            print 'One value editted'

    # fname = source.split('.')
    # extension = fname.pop()
    # fname = '.'.join(i for i in fname) + '.updated.' + extension
    wb2.save(target)

if __name__ == "__main__":
    p_model = "/afs/unity.ncsu.edu/users/b/bli6/temoa/temoa_model/ReferenceModel.py"
    p_data = [
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/noIGCC-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/noIGCC-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/IGCC-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/IGCC-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/COAL-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/COAL-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/noCOAL-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NC/noCOAL-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/CCS-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/CCS-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/noCCS-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/noCCS-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/IGCCS-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/IGCCS-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/noIGCCS-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/noIGCCS-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/NGCCS-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/NGCCS-noCP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/noNGCCS-CP",
    # "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/noNGCCS-noCP",
    "/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/NCupdated/rc_analysis",
    ]

    dummy_temoa_options = DummyTemoaConfig()
    dummy_temoa_options.config = None
    dummy_temoa_options.keepPyomoLP = False
    dummy_temoa_options.saveTEXTFILE = False
    dummy_temoa_options.path_to_db_io = None
    dummy_temoa_options.saveEXCEL = False

    # db file should be under the p_data directory
    # dummy_temoa_options.output = "NCreference.db"
    dummy_temoa_options.output = "NCupdated.db"

    # p_model = "/mnt/disk2/bli6/TEMOA_stochastic/Farmer/ReferenceModel.py"
    # p_data = "/mnt/disk2/bli6/TEMOA_stochastic/Farmer/scenariodata"

    # p_model = "/afs/unity.ncsu.edu/users/b/bli6/temoa/temoa_model/ReferenceModel.py"
    # p_data = "/afs/unity.ncsu.edu/users/b/bli6/temoa/tools/utopia_demand"

    # p_model = '/afs/unity.ncsu.edu/users/b/bli6/temoa/temoa_model/ReferenceModel.py'
    # p_data  = '/afs/unity.ncsu.edu/users/b/bli6/TEMOA_stochastic/test_twotechs_1'

    do_test(p_model, p_data, dummy_temoa_options)
    # do_test(p_model, p_data)

    # source = 'C:\\Users\\bli\\Downloads\\tmp\\TemoaS_lab\\NC\\noIGCC-noCP\\NCreferenceS.R.xlsx'
    # target = 'template.xlsx'
    # handle_excel(source, target)