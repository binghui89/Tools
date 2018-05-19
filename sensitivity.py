from pyomo.environ import *
from pyomo.core import Constraint
from pyomo.opt import SolverFactory
import sys, os, platform
from matplotlib import pyplot as plt
import numpy as np
from collections import OrderedDict, defaultdict
from time import time
from IPython import embed as IP

if platform.system() == 'Linux':
    sys.path.append("/afs/unity.ncsu.edu/users/b/bli6/temoa/temoa_model")
    sys.path.append('/opt/ibm/ILOG/CPLEX_Studio1263/cplex/python/2.6/x86-64_linux')
elif platform.system() == 'Windows':
    sys.path.append('C:\\Users\\bli\\GitHub\\Temoa\\temoa_model')
    sys.path.append('D:\\software\\cplex\\python\\2.7\\x64_win64')
elif platform.system() == 'Darwin':
    sys.path.append('/Users/bli/git/temoa/temoa_model')
    sys.path.append('/Users/bli/Applications/IBM/ILOG/CPLEX_Studio1263/cplex/python/2.7/x86-64_osx')
else:
    print 'Unrecognized system! Exiting...'
    sys.exit(0)

def return_Temoa_model():
    from temoa_model import temoa_create_model
    model = temoa_create_model()

    model.dual  = Suffix(direction=Suffix.IMPORT)
    model.rc    = Suffix(direction=Suffix.IMPORT)
    model.slack = Suffix(direction=Suffix.IMPORT)
    model.lrc   = Suffix(direction=Suffix.IMPORT)
    model.urc   = Suffix(direction=Suffix.IMPORT)
    return model

def return_Temoa_data(model, list_dat):
    data = DataPortal(model = model)
    for d in list_dat:
        data.load(filename=d)
    return data

def return_c_vector(block, unfixed):
    # Note that this function is adapted function collect_linear_terms defined
    # in pyomo/repn/collect.py.
    from pyutilib.misc import Bunch
    from pyomo.core.base import  Var, Constraint, Objective, maximize, minimize
    from pyomo.repn import generate_canonical_repn
    #
    # Variables are constraints of block
    # Constraints are unfixed variables of block and the parent model.
    #
    vnames = set()
    for (name, data) in block.component_map(Constraint, active=True).items():
        vnames.add((name, data.is_indexed()))
    cnames = set(unfixed)
    for (name, data) in block.component_map(Var, active=True).items():
        cnames.add((name, data.is_indexed()))
    #
    A = {}
    b_coef = {}
    c_rhs = {}
    c_sense = {}
    d_sense = None
    v_domain = {}
    #
    # Collect objective
    #
    for (oname, odata) in block.component_map(Objective, active=True).items():
        for ndx in odata:
            if odata[ndx].sense == maximize:
                o_terms = generate_canonical_repn(-1*odata[ndx].expr, compute_values=False)
                d_sense = minimize
            else:
                o_terms = generate_canonical_repn(odata[ndx].expr, compute_values=False)
                d_sense = maximize
            for i in range(len(o_terms.variables)):
                c_rhs[ o_terms.variables[i].parent_component().local_name, o_terms.variables[i].index() ] = o_terms.linear[i]
        # Stop after the first objective
        break
    return c_rhs

def validate_coef(c0, instance, target_tech, target_year):
    # This function validates if c0 equals the correct coefficient of process
    # (target_tech, target_year)
    t = target_tech
    v = target_year
    P_0 = min( instance.time_optimize )
    P_e = instance.time_future.last()
    GDR = value( instance.GlobalDiscountRate )
    MLL = instance.ModelLoanLife
    MPL = instance.ModelProcessLife
    LLN = instance.LifetimeLoanProcess
    x   = 1 + GDR    # convenience variable, nothing more.
    period_available = set()
    for p in instance.time_future:
        if (p, t, v) in instance.CostFixed.keys():
            period_available.add(p)
    c_i = ( 
            instance.CostInvest[t, v] 
            * instance.LoanAnnualize[t, v] 
            * ( LLN[t, v] if not GDR else 
                (x**(P_0 - v + 1) 
                * ( 1 - x **( -value(LLN[t, v]) ) ) 
                / GDR) 
                ) 
    ) * (
		  (
			  1 -  x**( -min( value(instance.LifetimeProcess[t, v]), P_e - v ) )
		  )
		  /(
			  1 -  x**( -value( instance.LifetimeProcess[t, v] ) ) 
		  )
    )

    c_f = sum( 
        instance.CostFixed[p, t, v]
        * ( MPL[p, t, v] if not GDR else
            (x**(P_0 - p + 1)
            * ( 1 - x**( -value(MPL[p, t, v]) ) )
            / GDR ) 
            )
        for p in period_available 
    ) 
    c = c_i + c_f

    if (c - c0) <= 1E-5:
        return True
    else:
        return False

def sensitivity(dat, techs):
    from temoa_model import temoa_create_model
    model = temoa_create_model()
    
    model.dual  = Suffix(direction=Suffix.IMPORT)
    model.rc    = Suffix(direction=Suffix.IMPORT)
    model.slack = Suffix(direction=Suffix.IMPORT)
    model.lrc   = Suffix(direction=Suffix.IMPORT)
    model.urc   = Suffix(direction=Suffix.IMPORT)

    data = DataPortal(model = model)
    for d in dat:
        data.load(filename=d)
    instance = model.create_instance(data)
    optimizer = SolverFactory('cplex')
    optimizer.options['lpmethod'] = 1 # Use primal simplex
    results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
    instance.solutions.load_from(results)

    coef_CAP = dict()
    scal_CAP = dict()
    # Break-even investment cost for this scenario, indexed by technology
    years    = list()
    bic_s    = dict()
    ic_s     = dict() # Raw investment costs for this scenario, indexed by tech
    cap_s    = dict()
    for t in techs:
        vintages = instance.vintage_optimize
        P_0 = min( instance.time_optimize )
        GDR = value( instance.GlobalDiscountRate )
        MLL = instance.ModelLoanLife
        MPL = instance.ModelProcessLife
        LLN = instance.LifetimeLoanProcess
        x   = 1 + GDR    # convenience variable, nothing more.

        bic_s[t] = list()
        ic_s[t]  = list()
        cap_s[t] = list()
        years = vintages.value
        for v in vintages:
            period_available = set()
            for p in instance.time_future:
                if (p, t, v) in instance.CostFixed.keys():
                    period_available.add(p)
            c_i = ( 
                    instance.CostInvest[t, v] 
                    * instance.LoanAnnualize[t, v] 
                    * ( LLN[t, v] if not GDR else 
                        (x**(P_0 - v + 1) 
                        * ( 1 - x **( -value(LLN[t, v]) ) ) 
                        / GDR) 
                      ) 
            )

            c_s = (-1)*(
                value( instance.CostInvest[t, v] )
                * value( instance.SalvageRate[t, v] )
                / ( 1 if not GDR else 
                    (1 + GDR)**( 
                        instance.time_future.last() 
                        - instance.time_future.first()
                        - 1
                        ) 
                    )
                )

            c_f = sum( 
                instance.CostFixed[p, t, v]
                * ( MPL[p, t, v] if not GDR else
                    (x**(P_0 - p + 1)
                    * ( 1 - x**( -value(MPL[p, t, v]) ) )
                    / GDR ) 
                  )
                for p in period_available 
            ) 

            c = c_i + c_s + c_f
            s = (c - instance.lrc[instance.V_Capacity[t, v]])/c
            coef_CAP[t, v] = c
            scal_CAP[t, v] = s # Must reduce TO this percentage
            bic_s[t].append(scal_CAP[t, v]*instance.CostInvest[t, v])
            ic_s[t].append(instance.CostInvest[t, v])
            cap_s[t].append( value( instance.V_Capacity[t, v] ) )

        # print "Tech\tVintage\tL. RC\tCoef\tU .RC\tScale\tBE IC\tBE FC\tIC\tFC\tCap"
        print "{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>7s}\t{:>5s}\t{:>3s}\t{:>5s}".format('Tech','Vintage', 'L. RC', 'Coef', 'U. RC', 'Scale', 'BE IC', 'BE FC', 'IC', 'FC', 'Cap')
        for v in vintages:
            lrc = instance.lrc[instance.V_Capacity[t, v]]
            urc = instance.urc[instance.V_Capacity[t, v]]

            # print "{:>s}\t{:>g}\t{:>.0f}\t{:>.0f}\t{:>.0f}\t{:>.3f}\t{:>.1f}\t{:>.1f}\t{:>.0f}\t{:>.0f}\t{:>.3f}".format(
            print "{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>7.1f}\t{:>5.0f}\t{:>3.0f}\t{:>5.3f}".format(
            t, v, lrc, coef_CAP[t, v], urc, scal_CAP[t, v], 
            scal_CAP[t, v]*instance.CostInvest[t, v], 
            scal_CAP[t, v]*instance.CostFixed[v, t, v], # Use the FC of the first period
            instance.CostInvest[t,v],
            instance.CostFixed[v, t, v],
            value(instance.V_Capacity[t, v])
            )

    IP()
    print 'Dual and slack variables for emission caps:'
    for e in instance.commodity_emissions:
        for p in instance.time_optimize:
            if (p, e) in instance.EmissionLimitConstraint:
                print p, e, instance.dual[instance.EmissionLimitConstraint[p, e]], '\t', instance.slack[instance.EmissionLimitConstraint[p, e]]
    return years, bic_s, ic_s

    print 'Dual and slack variables for Commodity Demand Constraints'
    for c in instance.commodity_demand:
        for p in instance.time_optimize:
            for s in instance.time_season:
                for tod in instance.time_of_day:
                    print p, s, tod, instance.dual[instance.DemandConstraint[p,s,tod,c]], instance.slack[instance.DemandConstraint[p,s,tod,c]]

def sensitivity_api(dat, techs, algorithm=None):
    import cplex
    model    = return_Temoa_model()
    data     = return_Temoa_data(model, dat)
    instance = model.create_instance(data)
    instance.write('tmp.lp', io_options={'symbolic_solver_labels':True})
    c = cplex.Cplex('tmp.lp')
    os.remove('tmp.lp')
    c.set_results_stream(None) # Turn screen output off

    msg = ''
    if algorithm:
        if algorithm == "o":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.auto)
        elif algorithm == "p":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.primal)
        elif algorithm == "d":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.dual)
        elif algorithm == "b":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
            c.parameters.barrier.crossover.set(
                c.parameters.barrier.crossover.values.none)
        elif algorithm == "h":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
        elif algorithm == "s":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.sifting)
        elif algorithm == "c":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.concurrent)
        else:
            raise ValueError(
                'method must be one of "o", "p", "d", "b", "h", "s" or "c"')

    try:
        c.solve()
    except CplexSolverError:
        print("Exception raised during solve")
        return

    vintages = list(instance.vintage_optimize)
    coef_CAP = dict()
    scal_CAP = dict()
    # Break-even investment cost for this scenario, indexed by technology
    years    = list()
    bic_s    = dict()
    ic_s     = dict() # Raw investment costs for this scenario, indexed by tech
    cap_s    = dict()
    clb_s    = dict()
    cub_s    = dict()
    for t in techs:
        bic_s[t] = list()
        ic_s[t]  = list()
        cap_s[t] = list()
        for v in vintages:
            target_var  = 'V_Capacity(' + t + '_' + str(v) + ')'
            c0          = c.objective.get_linear(target_var)
            clb, cub    = c.solution.sensitivity.objective(target_var) # Coefficient lower bound, coefficient upper bound
            if cub > 1E5:
                cub = 0 # Infinity
            clb_s[t, v], cub_s[t, v] = clb, cub
            if not validate_coef(c0, instance, t, v):
                print 'Error!'
                sys.exit(0)
            coef_CAP[t, v] = c0
            scal_CAP[t, v] = clb/c0 # Must reduce TO this percentage
            bic_s[t].append(scal_CAP[t, v]*instance.CostInvest[t, v])
            ic_s[t].append(instance.CostInvest[t, v])
            cap_s[t].append( c.solution.get_values(target_var) )

        print "{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>7s}\t{:>5s}\t{:>3s}\t{:>5s}".format('Tech','Vintage', 'L. CB', 'Coef', 'U. CB', 'Scale', 'BE IC', 'BE FC', 'IC', 'FC', 'Cap')
        msg += "{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>7s}\t{:>5s}\t{:>3s}\t{:>5s}".format('Tech','Vintage', 'L. CB', 'Coef', 'U. CB', 'Scale', 'BE IC', 'BE FC', 'IC', 'FC', 'Cap')
        msg == '\n'
        for v in vintages:
            print "{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>7.1f}\t{:>5.0f}\t{:>3.0f}\t{:>5.3f}".format(
            t,
            v, 
            clb_s[t, v],
            coef_CAP[t, v],
            cub_s[t, v],
            scal_CAP[t, v],
            scal_CAP[t, v]*instance.CostInvest[t, v], 
            scal_CAP[t, v]*instance.CostFixed[v, t, v], # Use the FC of the first period
            instance.CostInvest[t,v],
            instance.CostFixed[v, t, v],
            cap_s[t][vintages.index(v)]
            )

            msg += "{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>7.1f}\t{:>5.0f}\t{:>3.0f}\t{:>5.3f}".format(
            t,
            v, 
            clb_s[t, v],
            coef_CAP[t, v],
            cub_s[t, v],
            scal_CAP[t, v],
            scal_CAP[t, v]*instance.CostInvest[t, v], 
            scal_CAP[t, v]*instance.CostFixed[v, t, v], # Use the FC of the first period
            instance.CostInvest[t,v],
            instance.CostFixed[v, t, v],
            cap_s[t][vintages.index(v)]
            )
            msg += '\n'
    
    return msg

def bin_search(tech, vintage, dat, eps = 0.01, all_v = False):
    # Sensitivity analysis by binary search to find break-even scaling factor 
    # for a technology.
    # tech     -> Target technology.
    # vintage  -> Target vintage. It is break-even when capacity in this year >= 0 
    # dat      -> A list of .dat files.
    # eps      -> Convergence tolerance
    # all_v    -> A flag used indicate the costs of which vintages are subject 
    # to change. If it is FALSE, then only the investment costs and fixed costs
    # in the target vintage will be altered, otherwise all vintages are affected
    # Note that, only the capacity in the target vintage will be monitored and 
    # be used as the signal of break-even.
    monitor_year = vintage
    monitor_tech = tech

    t0 = time()
    time_mark = lambda: time() - t0 

    model = return_Temoa_model()
    optimizer = SolverFactory('cplex')
    data = return_Temoa_data(model, dat)
    instance = model.create_instance(data)

    time_optimize = [ i for i in data['time_future'] ]
    time_optimize.sort()
    ic0 = dict()
    fc0 = dict()
    if all_v:
        for v in time_optimize:
            if (monitor_tech, v) in data['CostInvest']:
                ic0[monitor_tech, v] = data['CostInvest'][monitor_tech, v]
                for p in time_optimize:
                    if (p, monitor_tech, v) in data['CostFixed']:
                        fc0[p, monitor_tech, v] = data['CostFixed'][p, monitor_tech, v]
    else:
        ic0[monitor_tech, monitor_year] = data['CostInvest'][monitor_tech, monitor_year]
        for p in time_optimize:
            if (p, monitor_tech, monitor_year) in data['CostFixed']:
                fc0[p, monitor_tech, monitor_year] = data['CostFixed'][p, monitor_tech, monitor_year]

    cap_target = 0
    scale_u = 1.0
    scale_l = 0.0

    history = dict()
    history['scale_u'] = [scale_u]
    history['scale_l'] = [scale_l]

    counter = 0
    scale_this = scale_u # Starting scale

    print 'Iteration # {} starts at {} s'.format( counter, time_mark() )
    instance = model.create_instance(data)
    instance.preprocess()
    results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
    instance.solutions.load_from(results)
    cap_target = value( instance.V_Capacity[monitor_tech, monitor_year] )
    print 'Iteration # {} solved at {} s'.format( counter, time_mark() )
    print 'Iteration # {}, scale: {:1.2f}, capacity: {} GW'.format( 
        counter,
        scale_this,
        cap_target
    )
    if 1.0 - scale_this <= eps and cap_target > 0:
        return scale_this

    while (scale_u - scale_l) >= eps and counter <= 20:
        if cap_target <= 0:
            scale_u = scale_this
            history['scale_u'].append(scale_u)
        else:
            scale_l = scale_this
            history['scale_l'].append(scale_l)
        counter += 1

        scale_this = (scale_u + scale_l)*0.5
        for k in ic0:
            data['CostInvest'][k] = scale_this*ic0[k]
        for k in fc0:
            data['CostFixed'][k] = scale_this*fc0[k]

        print 'Iteration # {} starts at {} s'.format( counter, time_mark() )
        instance = model.create_instance(data)
        instance.preprocess()
        results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
        instance.solutions.load_from(results)
        cap_target = value( instance.V_Capacity[monitor_tech, monitor_year] )
        print 'Iteration # {} solved at {} s'.format( counter, time_mark() )
        print 'Iteration # {}, scale: {:1.2f}, capacity: {} GW'.format( 
            counter,
            scale_this,
            cap_target)
    return (scale_u + scale_l)/2.0

def sen_range_api(tech, vintage, scales, list_dat):
    # This function is adapted from CPLEX's example script lpex2.py
    # It does the same thing as sen_range, but with CPLEX API for Python

    # Given a range of scaling factor for coefficient of a specific V_Capacity, 
    # returns objective value, reduced cost, capacity etc. for each scaling 
    # factor
    from openpyxl import Workbook
    import cplex
    target_year = vintage
    target_tech = tech
    target_var0 = 'V_Capacity(' + target_tech + '_' + str(target_year) + ')'
    algmap = {
        'primal simplex': 'p',
        'dual simplex':   'd',
        'barrier':        'h', # This is cross-over mode, since pure interior causes problems
        'default':        'o',
    } # cplex definition

    t0 = time()
    time_mark = lambda: time() - t0

    model = return_Temoa_model()
    data = return_Temoa_data(model, list_dat)
    instance = model.create_instance(data)

    ic0         = data['CostInvest'][target_tech, target_year]
    fc0         = data['CostFixed'][target_year, target_tech, target_year]
    all_periods = data['time_future']

    obj  = dict()
    cap  = dict()
    coef = dict()
    bic  = dict()
    bfc  = dict()
    ic   = dict() # Original IC
    fc   = dict() # Original FC
    clb  = dict() # Lower bound of objective coefficient
    cub  = dict() # Upper bound of objective coefficient
    rc   = dict() # Reduced cost

    for algorithm in ['barrier', 'dual simplex', 'primal simplex']:
        print 'Algorithm: {}'.format( algorithm )
        instance.write('tmp.lp', io_options={'symbolic_solver_labels':True})
        c = cplex.Cplex('tmp.lp')
        os.remove('tmp.lp')
        c.set_results_stream(None) # Turn screen output off
        c0 = c.objective.get_linear(target_var0)
        if not validate_coef(c0, instance, target_tech, target_year):
            print 'Error!'
            sys.exit(0)
        print '[{:>9.2f}] CPLEX model loaded.'.format( time_mark() )

        if algmap[algorithm] == "o":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.auto)
        elif algmap[algorithm] == "p":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.primal)
        elif algmap[algorithm] == "d":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.dual)
        elif algmap[algorithm] == "b":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
            c.parameters.barrier.crossover.set(
                c.parameters.barrier.crossover.values.none)
        elif algmap[algorithm] == "h":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
        elif algmap[algorithm] == "s":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.sifting)
        elif algmap[algorithm] == "c":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.concurrent)
        else:
            raise ValueError(
                'method must be one of "o", "p", "d", "b", "h", "s" or "c"')

        obj_alg  = list()
        cap_alg  = defaultdict(list)
        coef_alg = defaultdict(list)
        bic_alg  = defaultdict(list)
        bfc_alg  = defaultdict(list)
        ic_alg   = defaultdict(list)
        fc_alg   = defaultdict(list)
        clb_alg  = defaultdict(list)
        cub_alg  = defaultdict(list)
        rc_alg   = defaultdict(list)
        for s in scales:
            print '[{:>9.2f}] Scale: {:>.3f} starts'.format(time_mark(), s)
            c.objective.set_linear(target_var0, s*c0)

            try:
                c.solve()
            except CplexSolverError:
                print("Exception raised during solve")
                return

            obj_alg.append( c.solution.get_objective_value() )
            for y in instance.time_optimize:
                key = str(y)
                target_var   = 'V_Capacity(' + target_tech + '_' + key + ')'
                coefficient  = c.objective.get_linear(target_var)
                if y != target_year:
                    if not validate_coef(coefficient, instance, target_tech, y):
                        print 'Error!'
                        sys.exit(0)
                capacity = c.solution.get_values(target_var)
                try:
                    # Out of some unknow reason, sometimes this function will 
                    # fail even though the model is totally feasible.
                    # Notes: This function fails when cross-over is not selected
                    # when barrier algorithm is selected
                    c_bound = c.solution.sensitivity.objective(target_var)
                    s_be    = c_bound[0] / coefficient # Break-even scale
                except:
                    c_bound = [None, None]
                    s_be    = None
                cost_i = s*value( instance.CostInvest[target_tech, y] )
                cost_f = s*value( instance.CostFixed[y, target_tech, y] )
                

                cap_alg[key].append(capacity)
                coef_alg[key].append(coefficient)
                ic_alg[key].append(cost_i)
                fc_alg[key].append(cost_f)
                if s_be:
                    bic_alg[key].append(s_be*cost_i)
                    bfc_alg[key].append(s_be*cost_f)
                else:
                    bic_alg[key].append(None)
                    bfc_alg[key].append(None)
                clb_alg[key].append( c_bound[0] )
                cub_alg[key].append( c_bound[1] )
                rc_alg[key].append( c.solution.get_reduced_costs(target_var) )

        obj[algorithm]  = obj_alg
        cap[algorithm]  = cap_alg
        coef[algorithm] = coef_alg
        bic[algorithm]  = bic_alg
        bfc[algorithm]  = bfc_alg
        ic[algorithm]   = ic_alg
        fc[algorithm]   = fc_alg
        clb[algorithm]  = clb_alg
        cub[algorithm]  = cub_alg
        rc[algorithm]   = rc_alg

        # Write to Excel spreadsheet
        print '[{:>9.2f}] Saving to Excel spreadsheet'.format( time_mark() )
        row_title = [
            'scale',       'obj',       'cap', 'clb', 'coef', 
            'cub',   'bic (clb)', 'bfc (clb)',  'ic',   'fc',
            'rc'
        ]
        wb = Workbook()
        # for ws_title in cap_alg:
        for year in all_periods:
            ws_title = str(year)
            if ws_title not in cap_alg:
                continue
            ws = wb.create_sheet(ws_title)

            row = [
                scales, 
                obj_alg, 
                cap_alg[ws_title], 
                clb_alg[ws_title], 
                coef_alg[ws_title], 
                cub_alg[ws_title], 
                bic_alg[ws_title], 
                bfc_alg[ws_title], 
                ic_alg[ws_title], 
                fc_alg[ws_title],
                rc_alg[ws_title]
            ]

            # Note Python starts from 0, but row number starts from 1
            for j in range(0, len(row_title) ):
                cell = ws.cell(row = 1, column = j + 1)
                cell.value = row_title[j]
            for i in range(0, len(scales)):
                for j in range(0, len(row_title)):
                    cell = ws.cell(row = i + 2, column = j + 1)
                    cell.value = row[j][i]
        fname = '.'.join(
            [target_tech, str(target_year)]
            + [ i[:-4] for i in list_dat ] # Remove the .dat extension
            + [algorithm]
        ) # tech_name.year.dat_file_name.algorithm.xlsx
        wb.save(fname + '.xlsx')

def sen_range(tech, vintage, scales, dat):
    # Given a range of scaling factor for coefficient of a specific V_Capacity, 
    # returns objective value, reduced cost, capacity etc. for each scaling 
    # factor
    from openpyxl import Workbook
    target_year = vintage
    target_tech = tech
    algmap = {
        'primal simplex': 1,
        'dual simplex':   2,
        'barrier':        4,
        'default':        0,
    } # cplex definition

    t0 = time()
    time_mark = lambda: time() - t0

    model = return_Temoa_model()
    data = return_Temoa_data(model, dat)
    optimizer = SolverFactory('cplex')

    ic0         = data['CostInvest'][target_tech, target_year]
    fc0         = data['CostFixed'][target_year, target_tech, target_year]
    all_periods = data['time_future']

    obj  = dict()
    cap  = dict()
    lrc  = dict()
    coef = dict()
    urc  = dict()
    bic  = dict()
    bfc  = dict()
    ic   = dict() # Original IC
    fc   = dict() # Original FC

    for algorithm in ['barrier', 'dual simplex', 'primal simplex']:
        optimizer.options['lpmethod'] = algmap[algorithm]
        print 'Algorithm: {}'.format( algorithm )

        obj_alg  = list()
        cap_alg  = defaultdict(list)
        lrc_alg  = defaultdict(list)
        coef_alg = defaultdict(list)
        urc_alg  = defaultdict(list)
        bic_alg  = defaultdict(list)
        bfc_alg  = defaultdict(list)
        ic_alg   = defaultdict(list)
        fc_alg   = defaultdict(list)
        for s in scales:
            print '[{:>9.2f}] Scale: {:>.3f} starts'.format(time_mark(), s)
            data['CostInvest'][target_tech, target_year] = s*ic0
            for y in data['time_future']:
                if (y, target_tech, target_year) in data['CostFixed']:
                    data['CostFixed'][y, target_tech, target_year] = s*fc0
            instance = model.create_instance(data)
            instance.preprocess()
            results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
            instance.solutions.load_from(results)

            obj_alg.append( value(instance.TotalCost) )
            for y in instance.time_optimize:
                key = str(y)
                c_vector = return_c_vector(instance, [])
                coefficient = c_vector[ ( 'V_Capacity', (target_tech, y) )]
                capacity = value(instance.V_Capacity[target_tech, y])
                lower_rc = value(
                    instance.lrc[ instance.V_Capacity[target_tech, y] ]
                )
                upper_rc = value(
                    instance.urc[ instance.V_Capacity[target_tech, y] ]
                )
                cost_i   = value( instance.CostInvest[target_tech, y] )
                cost_f   = value( instance.CostFixed[y, target_tech, y] )
                s_be = ( coefficient - lower_rc ) / coefficient # Break-even scale

                cap_alg[key].append(capacity)
                lrc_alg[key].append(lower_rc)
                coef_alg[key].append(coefficient)
                urc_alg[key].append(upper_rc)
                ic_alg[key].append(cost_i)
                fc_alg[key].append(cost_f)
                bic_alg[key].append(s_be*cost_i)
                bfc_alg[key].append(s_be*cost_f)

        obj[algorithm]  = obj_alg
        cap[algorithm]  = cap_alg
        lrc[algorithm]  = lrc_alg
        coef[algorithm] = coef_alg
        urc[algorithm]  = urc_alg
        bic[algorithm]  = bic_alg
        bfc[algorithm]  = bfc_alg
        ic[algorithm]   = ic_alg
        fc[algorithm]   = fc_alg

        # Write to Excel spreadsheet
        print '[{:>9.2f}] Saving to Excel spreadsheet'.format( time_mark() )
        row_title = [
            'scale', 'obj', 'cap', 'lrc', 'coef', 
            'urc',   'bic', 'bfc', 'ic',  'fc'
        ]
        wb = Workbook()
        # for ws_title in cap_alg:
        for year in all_periods:
            ws_title = str(year)
            if ws_title not in cap_alg:
                continue
            ws = wb.create_sheet(ws_title)

            row = [
                scales, 
                obj_alg, 
                cap_alg[ws_title], 
                lrc_alg[ws_title], 
                coef_alg[ws_title], 
                urc_alg[ws_title], 
                bic_alg[ws_title], 
                bfc_alg[ws_title], 
                ic_alg[ws_title], 
                fc_alg[ws_title]
            ]

            # Note Python starts from 0, but row number starts from 1
            for j in range(0, len(row_title) ):
                c = ws.cell(row = 1, column = j + 1)
                c.value = row_title[j]
            for i in range(0, len(scales)):
                for j in range(0, len(row_title)):
                    c = ws.cell(row = i + 2, column = j + 1)
                    c.value = row[j][i]
        fname = '.'.join(
            [target_tech, str(target_year)]
            + [ i[:-4] for i in dat ] # Remove the .dat extension
            + [algorithm]
        ) # tech_name.year.dat_file_name.algorithm.xlsx
        wb.save(fname + '.xlsx')

def explore_Cost_marginal(dat):
    from temoa_model import temoa_create_model
    model = temoa_create_model()
    
    model.dual  = Suffix(direction=Suffix.IMPORT)
    model.rc    = Suffix(direction=Suffix.IMPORT)
    model.slack = Suffix(direction=Suffix.IMPORT)
    model.lrc   = Suffix(direction=Suffix.IMPORT)
    model.urc   = Suffix(direction=Suffix.IMPORT)

    data = DataPortal(model = model)
    for d in dat:
        data.load(filename=d)
    instance = model.create_instance(data)

    # Deactivate the DemandActivity constraint
    # instance.DemandActivityConstraint.deactivate()
    # instance.preprocess()

    optimizer = SolverFactory('cplex')
    results = optimizer.solve(
        instance, 
        keepfiles=True,
        suffixes=['dual', 'urc', 'slack', 'lrc']
        )
    instance.solutions.load_from(results)

    print 'Dual and slack variables for emission caps:'
    for e in instance.commodity_emissions:
        for p in instance.time_optimize:
            if (p, e) in instance.EmissionLimitConstraint:
                print p, e, instance.dual[instance.EmissionLimitConstraint[p, e]], '\t', instance.slack[instance.EmissionLimitConstraint[p, e]]

    print 'Dual and slack variables for Commodity Demand Constraints'
    for c in instance.commodity_demand:
        for p in instance.time_optimize:
            for s in instance.time_season:
                for tod in instance.time_of_day:
                    print p, s, tod, instance.dual[instance.DemandConstraint[p,s,tod,c]], instance.slack[instance.DemandConstraint[p,s,tod,c]]

def LC_calculate(dat):

    # Electricity generating technologies
    tech_gen = [
        'IMPELCNGCEA',
        'IMPELCNGAEA',
        'IMPELCDSLEA',
        'IMPURNA',
        'IMPELCBIGCCEA',
        'IMPELCBIOSTM',
        'IMPELCGEO',
        'IMPSOL',
        'IMPWND',
        'IMPELCHYD',
        'IMPLFGICEEA',
        'IMPLFGGTREA',
        'IMPELCCOAB',

        'ENGACC05',
        'ENGACT05',
        'ENGAACC',
        'ENGAACT',
        'ENGACCCCS',
        'ENGACCR',
        'ENGACTR',
        'ECOALSTM',
        'ECOALIGCC',
        'ECOALIGCCS',
        'ECOASTMR',
        'EDSLCTR',
        'EURNALWR',
        'EURNALWR15',
        'EBIOIGCC',
        'EBIOSTMR',        
        'EGEOBCFS',        
        'ESOLPVCEN',       
        'ESOLSTCEN',       
        'ESOLPVR',         
        'EWNDON',          
        'EWNDOFS',         
        'EWNDOFD',         
        'EHYDCONR',        
        'EHYDREVR',        
        'ELFGICER',        
        'ELFGGTR',         
        'EHYDGS']
    
    tech_misc = [
        'E_BLND_BITSUBLIG_COALSTM_R',   # blending tech to collect bit subbit and lig coal for existing coal steam plant');
        'E_BLND_BIT_COALSTM_R',         # blending tech to collect bit coal for existing coal steam plant');
        'E_PTNOXSCR_COAB',              # nox passthrough tech for bituminous coal after LNB retrofit or passthrough and before existing coal steam plant');
        'E_PTNOXLNB_COAB',              # nox passthrough tech for bituminous coal after so2 or co2 passthrough and before SCR or SNCR or passthrough');
        'E_PTCO2_COAB',                 # co2 passthrough tech for bituminous coal after FGD and before LNB');
        'E_EA_COAB',                    # co2 emission accounting tech for coal bituminous');
        'E_PTSO2_COABH',                # passthrough tech with no so2 removal from bit high sulfur before existing coal steam plant');
        'E_PTSO2_COABM',                # passthrough tech with no so2 removal from bit medium sulfur before existing coal steam plant');
        'E_PTSO2_COABL',                # passthrough tech with no so2 removal from bit low sulfur before existing coal steam plant');
        'E_BLND_BITHML_COALIGCC_N',     # blending tech to collect high medium low sulfur bit coal for new coal IGCC plant');
        'E_BLND_BITSUBLIG_COALSTM_N',   # blending tech to collect bit subbit and lig coal for new coal steam plant');
        'E_BLND_BITHML_COALSTM_N',      # blending tech to collect high medium low sulfur bit coal for new coal steam plant');
        'ELC2DMD'
        ]
    
    # NOx removal technologies
    tech_NOx = [
        'E_LNBSNCR_COAB_R', # Existing LNB w/SNCR retrofit for nox removal from BIT before existing coal STM');
        'E_LNBSNCR_COAB_N', # new LNB combined with SNCR retrofit for nox removal from bituminous before existing coal steam plant');
        'E_LNBSCR_COAB_R',  # existing LNB combined with SCR retrofit for nox removal from bituminous before existing coal steam plant');
        'E_LNBSCR_COAB_N',  # new LNB combined with SCR retrofit for nox removal from bituminous before existing coal steam plant');
        'E_SNCR_COAB_R',    # existing SNCR retrofit for nox removal from bituminous before existing coal steam plant');
        'E_SNCR_COAB_N',    # new SNCR retrofit for nox removal from bituminous before existing coal steam plant');
        'E_SCR_COAB_N',     # new SCR retrofit for nox removal from bituminous before existing coal steam plant');
        'E_LNB_COAB_R',     # existing LNB retrofit tech for nox removal from bituminous before existing coal steam plant');
        'E_LNB_COAB_N'     # new LNB retrofit tech for nox removal from bituminous before existing coal steam plant
        ]
    
    # SO2 removal technologies
    tech_SO2 = [
        'E_FGD_COABH_N', # new FGD retrofit tech for so2 removal from bit high sulfur before existing coal steam plant');
        'E_FGD_COABM_R', # existing FGD retrofit tech for so2 removal from bit medium sulfur before existing coal steam plant');
        'E_FGD_COABM_N', # new FGD retrofit tech for so2 removal from bit medium sulfur before existing coal steam plant');
        'E_FGD_COABL_R', # existing FGD retrofit tech for so2 removal from bit low sulfur before existing coal steam plant');
        'E_FGD_COABL_N'  # new FGD retrofit tech for so2 removal from bit low sulfur before existing coal steam plant');
        ]
    
    # CO2 removal technologies
    tech_CO2 = [
        'E_CCR_COAB',                   # co2 capture retrofit tech for bituminous coal to existing power plant located after FGD or passthrough and before LNB');
        'E_CCR_COALIGCC_N',             # co2 capture retrofit tech before coal IGCC plant');
        'E_CCR_COALSTM_N',              # co2 capture retrofit tech before new coal steam plant');
        'E_BLND_BITSUBLIG_COALIGCC_N'   # blending tech to collect bit subbit and lig coal for new coal IGCC plant');
        ]

    data = DataPortal(model = model)
    data.load(filename=dat)
    instance = model.create_instance(data)
    optimizer = SolverFactory('cplex')
    results = optimizer.solve(instance, suffixes=['dual', 'rc', 'slack', 'lrc'])
    instance.solutions.load_from(results)

    m = instance
    TAE  = dict()
    # Cost format: [Gen, NOx, SO2, CO2, misc] where misc supposed to be 0, 
    # otherwise there is something wrong with the cost classification table.
    TAIC = dict() # Total annualized investment cost
    TAFC = dict() # Total annualized fixed O&M cost
    TAVC = dict() # Total annualized variable O&M cost
    TAMC = dict() 
    TAC  = dict() # Total annualized cost
    
    LCOE = dict() # Levelized cost of energy
    LCOI = dict() # Levelized cost of investment
    LCOF = dict() # Levelized cost of fixed O&M
    LCOV = dict() # Levelized cost of variable O&M

    for year in m.time_optimize:
        year_str = str(year)

        # Annual energy output
        TAE[year_str] = sum(
            1/3.6 # Convert PJ to TWh
            *value( m.V_ActivityByPeriodTechAndOutput[S_p, S_t, S_o] )
            for S_p, S_t, S_o in m.V_ActivityByPeriodTechAndOutput.iterkeys()
            if S_p == year and S_o == 'ELC'
            )
        
        # Annualized investment cost
        TAIC[year_str] = [0, 0, 0, 0, 0] 
        TAFC[year_str] = [0, 0, 0, 0, 0] 
        TAVC[year_str] = [0, 0, 0, 0, 0] 
        for S_t, S_v in m.CostInvest.sparse_iterkeys():
            if S_v == year:
                if S_t in tech_gen:
                    TAIC[year_str][0] += (
                        value( m.CostInvest[S_t, S_v] )
                        *value( m.LoanAnnualize[S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                elif S_t in tech_NOx:
                    TAIC[year_str][1] += (
                        value( m.CostInvest[S_t, S_v] )
                        *value( m.LoanAnnualize[S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                elif S_t in tech_SO2:
                    TAIC[year_str][2] += (
                        value( m.CostInvest[S_t, S_v] )
                        *value( m.LoanAnnualize[S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                elif S_t in tech_CO2:
                    TAIC[year_str][3] += (
                        value( m.CostInvest[S_t, S_v] )
                        *value( m.LoanAnnualize[S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                else:
                    TAIC[year_str][4] += (
                        value( m.CostInvest[S_t, S_v] )
                        *value( m.LoanAnnualize[S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
            else:
                continue
        
        # Annualized fixed cost
        for S_p, S_t, S_v in m.CostFixed.sparse_iterkeys():
            if S_p == year:
                if S_t in tech_gen:
                    TAFC[year_str][0] += (
                        value( m.CostFixed[S_p, S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                elif S_t in tech_NOx:
                    TAFC[year_str][1] += (
                        value( m.CostFixed[S_p, S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                elif S_t in tech_SO2:
                    TAFC[year_str][2] += (
                        value( m.CostFixed[S_p, S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                elif S_t in tech_CO2:
                    TAFC[year_str][3] += (
                        value( m.CostFixed[S_p, S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
                else:
                    TAFC[year_str][4] += (
                        value( m.CostFixed[S_p, S_t, S_v] )
                        *value( m.V_Capacity[S_t, S_v] )
                    )
            else:
                continue
        
        # Annualized variable cost
        for S_p, S_t, S_v in m.CostVariable.sparse_iterkeys():
            if S_p == year:
                if S_t in tech_gen:
                    TAVC[year_str][0] += (
                        value( m.CostVariable[S_p, S_t, S_v] )
                        *value( m.V_ActivityByPeriodAndProcess[S_p, S_t, S_v] )
                    )
                elif S_t in tech_NOx:
                    TAVC[year_str][1] += (
                        value( m.CostVariable[S_p, S_t, S_v] )
                        *value( m.V_ActivityByPeriodAndProcess[S_p, S_t, S_v] )
                    )
                elif S_t in tech_SO2:
                    TAVC[year_str][2] += (
                        value( m.CostVariable[S_p, S_t, S_v] )
                        *value( m.V_ActivityByPeriodAndProcess[S_p, S_t, S_v] )
                    )
                elif S_t in tech_CO2:
                    TAVC[year_str][3] += (
                        value( m.CostVariable[S_p, S_t, S_v] )
                        *value( m.V_ActivityByPeriodAndProcess[S_p, S_t, S_v] )
                    )
                else:
                    TAVC[year_str][4] += (
                        value( m.CostVariable[S_p, S_t, S_v] )
                        *value( m.V_ActivityByPeriodAndProcess[S_p, S_t, S_v] )
                    )
            else:
                continue

        TAC[year_str] = [
            TAIC[year_str][i] + 
            TAFC[year_str][i] + 
            TAVC[year_str][i] for i in range( 0, len(TAIC[year_str]) )
        ]
        LCOE[year_str] = [ i/TAE[year_str] for i in TAC[year_str] ]

        LCOI[year_str] = [ i/TAE[year_str] for i in TAIC[year_str] ]
        LCOF[year_str] = [ i/TAE[year_str] for i in TAFC[year_str] ]
        LCOV[year_str] = [ i/TAE[year_str] for i in TAVC[year_str] ]
        
        # print year_str, LCOE[year_str], sum(LCOI[year_str]), sum(LCOF[year_str]), sum(LCOV[year_str])

    return LCOI, LCOF, LCOV 

def do_LCcalculate():
    # print "Reference scenario"
    # dat = 'sql/reference.dat'
    # LC_calculate(dat)
    # print "High oil scenario"
    # dat = 'sql/reference.high_oil.dat'
    # LC_calculate(dat)
    # print "Low oil scenario"
    # dat = 'sql/reference.high_res.dat'
    # LC_calculate(dat)

    # print "CPP scenario"
    # dat = 'sql/cpp.dat'
    # LC_calculate(dat)
    # print "CPP + high oil scenario"
    # dat = 'sql/cpp.high_oil.dat'
    # LC_calculate(dat)
    # print "CPP + low oil scenario"
    # dat = 'sql/cpp.high_res.dat'
    # LC_calculate(dat)

    print "Test scenario"
    dat = 'sql20151124.test.dat'
    LC_calculate(dat)

def plot_LCOE():
    periods = range(2015, 2055, 5)
    LCOI = dict()
    LCOF = dict()
    LCOV = dict()
    scenarios = {'R':     'sql20151124/reference.dat',
                 'LF':    'sql20151124/reference.high_res.dat',
                 'HF':    'sql20151124/reference.high_oil.dat',
                 'CPP':   'sql20151124/cpp.dat',
                 'CPPLF': 'sql20151124/cpp.high_res.dat',
                 'CPPHF': 'sql20151124/cpp.high_oil.dat'
                }
    for s in scenarios.keys():
        print "Scenario: ", s
        LCOI_split, LCOF_split, LCOV_split = LC_calculate(scenarios[s])
        LCOI[s] = [sum(LCOI_split[str(p)]) for p in periods]
        LCOF[s] = [sum(LCOF_split[str(p)]) for p in periods]
        LCOV[s] = [sum(LCOV_split[str(p)]) for p in periods]

    bar_width = 0.7
    left_position = np.array(range(2015, 2055, 5))
    # c_ref = [0.1, 0.1, 0.1]
    # c_cpp = 'g'
    colors = {
        'LF':    [1, 1, 1],
        'R':     [1, 1, 1],
        'HF':    [1, 1, 1],
        'CPPLF': [0, 1, 0],
        'CPP':   [0, 1, 0],
        'CPPHF': [0, 1, 0]
    }
    hatchs = {
        'LF':    '/',
        'R':     None,
        'HF':    '\\',
        'CPPLF': '/',
        'CPP':   None,
        'CPPHF': '\\'
    }
    handles = list()

    for s in ['LF', 'R', 'HF', 'CPPLF', 'CPP', 'CPPHF']:
        h = plt.bar(left_position, np.array(LCOV[s]), 
                    bar_width, 
                    # alpha=0.3, 
                    color=[0.9*i for i in colors[s]],
                    hatch=hatchs[s])
        handles.append(h)
        h = plt.bar(left_position, np.array(LCOF[s]),
                    bar_width, 
                    bottom=np.array(LCOV[s]),
                    # alpha=0.5, 
                    color=[0.5*i for i in colors[s]],
                    hatch=hatchs[s])
        handles.append(h)
        h = plt.bar(left_position, np.array(LCOI[s]),
                    bar_width, 
                    bottom=np.array(LCOV[s])+np.array(LCOF[s]),
                    # alpha=1, 
                    color=[0.2*i for i in colors[s]],
                    hatch=hatchs[s])
        handles.append(h)
        left_position = left_position + bar_width

    plt.xlabel('Periods')
    plt.ylabel('LCOE ($/MWh)')
    plt.xticks(left_position-3*bar_width, [str(p) for p in range(2015, 2055, 5)])
    plt.show()

def plot_breakeven(years, bic, ic):
    # bic is a dictionary, ic is a list of the raw investment costs
    # ic = [x, x, ..., x], the length of which equals to the length of years
    # bic[scenario] = [x, x, x... x] where the length equals to the number 
    # of optimized periods.
    sen_color_map = {
        'IC':    [0.9, 0.9, 0.9],
        'LF':    'black',
        'R':     'black',
        'HF':    'black',
        'HD':    'black',
        'CPPLF': 'green',
        'CPP':   'green',
        'CPPHF': 'green',
        'CPPHD': 'green'
    }

    sen_lstyle_map = {
        'IC':    None,
        'LF':    '--',
        'R':     '-',
        'HF':    '-.',
        'HD':    ':',
        'CPPLF': '--',
        'CPP':   '-',
        'CPPHF': '-.',
        'CPPHD': ':'
    }

    sen_marker_map = {
        'IC':    None,
        'LF':    's',
        'R':     's',
        'HF':    's',
        'HD':    's',
        'CPPLF': 's',
        'CPP':   's',
        'CPPHF': 's',
        'CPPHD': 's'
    }

    scenarios = bic.keys()
    h = plt.figure()
    ax = plt.subplot(111)
    ax.fill_between(years, 0, ic, 
        facecolor = sen_color_map['IC']
        )

    for s in scenarios:
        ax.plot(years, bic[s], 
            color = sen_color_map[s],
            # marker = sen_marker_map[s],
            linestyle = sen_lstyle_map[s]
            )
    ax.yaxis.grid(True)
    plt.ylabel('$/MWh')
    plt.xlim( ( years[0]-5, years[-1]+5 ) )
    return ax

def bin_search_and_range():
    def return_range(bs):
        # Given break-even scaling factor, return an appropriate range
        ub = None
        lb = None
        if bs >= 0.85:
            lb = int(bs*100) - 15
            ub = 100
        elif bs <= 0.15:
            lb = 1
            ub = int(bs*100) + 15
        else:
            lb = int(bs*100) - 15
            ub = int(bs*100) + 15
        return [0.001*i for i in range(lb*10, ub*10, 10)]

    list_file = ['reference.dat', 'NCupdated_noLeadTime.dat']
    list_tech = ['EBIOIGCC', 'EURNALWR15']
    monitor_vintage = 2020
    eps = 0.01
    for f in list_file:
        for t in list_tech:
            bs = bin_search(t, monitor_vintage, [f], eps)
            sen_range( t, monitor_vintage, return_range(bs), [f] )

if __name__ == "__main__":
    # sen_bin_search(
    #     'ECOALIGCCS', 
    #     2020,
    #     ['reference.dat'],
    #     0.01
    # )
    scales = [0.001* i for i in range(250, 260, 10)]
    sen_range(
        'ECOALIGCCS', 
        2020,
        scales,
        ['reference.dat']
    )
    # do_sensitivity_new()
    # do_sensitivity_old()
    # explore_Cost_marginal(['/afs/unity.ncsu.edu/users/b/bli6/TEMOA_NC/sql20170417/results/R/NCreference.R.dat'])
